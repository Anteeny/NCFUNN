import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def create_ncf_database():
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Navy
    header_font = Font(name="Segoe UI", size=11, bold=True, color="F59E0B") # Gold
    title_font = Font(name="Segoe UI", size=14, bold=True, color="1E293B")
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    normal_font = Font(name="Segoe UI", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    # 1. MEMBERSHIP DASHBOARD SHEET
    ws_members = wb.create_sheet(title="Membership Dashboard")
    ws_members.views.sheetView[0].showGridLines = True
    
    member_headers = [
        "S/N", "First Name", "Surname", "Phone Number", "Email", "Gender", "Marital Status",
        "Occupation", "Residential Address", "Department", "Believers' Circle", "Cell",
        "G12 Leader", "DCA Basic", "Encounter", "DLI Basic", "Membership Level", "Attendance Status"
    ]
    
    ws_members.append(member_headers)
    for col_num in range(1, len(member_headers) + 1):
        cell = ws_members.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    sample_members = [
        [1, "Mmesoma", "Ozor", "09029433375", "mmesoma@ncfunn.org", "Female", "Single", "Student", "Ikejani Staff Quarters", "Pastorate", "Zion BC", "Franco Cell 1", "P.Ngozi Okeke", "Yes", "Yes", "Yes", "Pastor", "=IFS(COUNTIF(Attendance!D2:K2, 1)>=6, \"Regular Attender\", COUNTIF(Attendance!D2:K2, 1)>=3, \"Drifting Attender\", COUNTIF(Attendance!D2:K2, 1)>=1, \"Occasional Attender\", TRUE, \"Inactive\")"],
        [2, "Royal-Priest", "Eroh", "08076229301", "royal@ncfunn.org", "Male", "Single", "Student", "Hilltop", "Pastorate", "Covenant BC", "Hilltop Cell 2", "Tony Ubagu", "Yes", "Yes", "Yes", "Leader", "=IFS(COUNTIF(Attendance!D3:K3, 1)>=6, \"Regular Attender\", COUNTIF(Attendance!D3:K3, 1)>=3, \"Drifting Attender\", COUNTIF(Attendance!D3:K3, 1)>=1, \"Occasional Attender\", TRUE, \"Inactive\")"],
        [3, "Ebuka", "Aziekwe", "08060883561", "ebuka@ncfunn.org", "Male", "Single", "Graduate", "Staff Quarters", "Pastorate", "Grace BC", "HQ Cell 1", "Tony Ubagu", "Yes", "Yes", "No", "Worker", "=IFS(COUNTIF(Attendance!D4:K4, 1)>=6, \"Regular Attender\", COUNTIF(Attendance!D4:K4, 1)>=3, \"Drifting Attender\", COUNTIF(Attendance!D4:K4, 1)>=1, \"Occasional Attender\", TRUE, \"Inactive\")"],
        [4, "Chiaza", "Ugwu", "08144199545", "chiaza@ncfunn.org", "Female", "Single", "Student", "Franco", "MVP", "Unassigned", "Unassigned", "Unassigned", "No", "No", "No", "MVP", "=IFS(COUNTIF(Attendance!D5:K5, 1)>=6, \"Regular Attender\", COUNTIF(Attendance!D5:K5, 1)>=3, \"Drifting Attender\", COUNTIF(Attendance!D5:K5, 1)>=1, \"Occasional Attender\", TRUE, \"Inactive\")"]
    ]

    for row_idx, row_data in enumerate(sample_members, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_members.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            if col_idx in [1, 6, 7, 14, 15, 16, 17, 18]:
                cell.alignment = Alignment(horizontal="center")

    # 2. CONTACTS SHEET
    ws_contacts = wb.create_sheet(title="Contacts")
    ws_contacts.views.sheetView[0].showGridLines = True
    
    contact_headers = ["S/N", "First Name", "Surname", "Phone Number", "Status"]
    ws_contacts.append(contact_headers)
    for col_num in range(1, len(contact_headers) + 1):
        cell = ws_contacts.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    sample_contacts = [
        [1, "Peter", "Okonkwo", "08121179178", "Contact"],
        [2, "Ebuka", "Eze", "07038172295", "Contact"],
        [3, "David", "Nwosu", "09032162359", "Contact"],
        [4, "Chiaza", "Ugwu", "08144199545", "MVP"],
        [5, "Chinonso", "Anayo", "07071140698", "Contact"]
    ]

    for row_idx, row_data in enumerate(sample_contacts, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_contacts.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            if col_idx in [1, 5]:
                cell.alignment = Alignment(horizontal="center")

    # Data Validation for Contacts Status
    dv_status = DataValidation(type="list", formula1='"Contact,MVP"', allow_blank=True)
    ws_contacts.add_data_validation(dv_status)
    dv_status.add("E2:E500")

    # 3. ATTENDANCE SHEET
    ws_att = wb.create_sheet(title="Attendance")
    ws_att.views.sheetView[0].showGridLines = True
    
    att_headers = ["S/N", "Full Name", "Phone", "Sunday 1", "Sunday 2", "Sunday 3", "Sunday 4", "Sunday 5", "Sunday 6", "Sunday 7", "Sunday 8", "Midweek 1", "Midweek 2"]
    ws_att.append(att_headers)
    for col_num in range(1, len(att_headers) + 1):
        cell = ws_att.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    sample_att = [
        [1, "Mmesoma Ozor", "09029433375", 1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
        [2, "Royal-Priest Eroh", "08076229301", 1, 1, 1, 0, 1, 1, 1, 0, 1, 1],
        [3, "Ebuka Aziekwe", "08060883561", 1, 0, 1, 0, 1, 0, 0, 1, 0, 1],
        [4, "Chiaza Ugwu", "08144199545", 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    ]

    for row_idx, row_data in enumerate(sample_att, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_att.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            if col_idx in [1] or col_idx >= 4:
                cell.alignment = Alignment(horizontal="center")

    # 4. SUMMARY SHEET
    ws_sum = wb.create_sheet(title="Summary")
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.cell(row=1, column=1, value="NCF DATABASE — EXECUTIVE SUMMARY KPI DASHBOARD").font = title_font
    
    kpi_headers = ["Metric Category", "Description / Rollup Formula", "Count"]
    for col_idx, h in enumerate(kpi_headers, start=1):
        cell = ws_sum.cell(row=3, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    kpis = [
        ["Total Prospects Logged", "Outreach contacts not yet MVPs", '=COUNTIF(Contacts!E:E, "Contact")'],
        ["Total MVPs (First Timers)", "First timers from beginning", '=COUNTIF(\'Membership Dashboard\'!Q:Q, "MVP")'],
        ["Regular Members", "Active church members", '=COUNTIF(\'Membership Dashboard\'!Q:Q, "Member")'],
        ["Workers & Volunteers", "Departmental workforce", '=COUNTIF(\'Membership Dashboard\'!Q:Q, "Worker")'],
        ["Discipleship Leaders", "G12 & Department Heads", '=COUNTIF(\'Membership Dashboard\'!Q:Q, "Leader")'],
        ["Pastors & Pastoral Care", "Pastorate council", '=COUNTIF(\'Membership Dashboard\'!Q:Q, "Pastor")'],
        ["Regular Attenders (>=75%)", "Scanned 8-Sunday Window", '=COUNTIF(\'Membership Dashboard\'!R:R, "Regular Attender")'],
        ["Drifting Attenders (37.5%-62.5%)", "Needs pastoral call", '=COUNTIF(\'Membership Dashboard\'!R:R, "Drifting Attender")'],
        ["Occasional Attenders (12.5%-25%)", "High risk of dropout", '=COUNTIF(\'Membership Dashboard\'!R:R, "Occasional Attender")'],
        ["Inactive Members (0%)", "Requires immediate follow-up", '=COUNTIF(\'Membership Dashboard\'!R:R, "Inactive")']
    ]

    for idx, (cat, desc, formula) in enumerate(kpis, start=4):
        c1 = ws_sum.cell(row=idx, column=1, value=cat)
        c2 = ws_sum.cell(row=idx, column=2, value=desc)
        c3 = ws_sum.cell(row=idx, column=3, value=formula)
        for c in [c1, c2, c3]:
            c.font = normal_font
            c.border = thin_border
        c3.font = bold_font
        c3.alignment = Alignment(horizontal="center")

    # 5. CELLS BCs SHEET
    ws_cells = wb.create_sheet(title="Cells BCs")
    ws_cells.views.sheetView[0].showGridLines = True
    
    cell_headers = ["S/N", "Zone", "Believers' Circle", "Cell Name", "Assigned Leader", "Meeting Location & Schedule"]
    ws_cells.append(cell_headers)
    for col_num in range(1, len(cell_headers) + 1):
        cell = ws_cells.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    sample_cells = [
        [1, "HQ Zone", "Covenant Springs BC", "Covenant Cell 1", "Royal-Priest Eroh", "HQ Main Auditorium Annex (Saturdays 5:00 PM)"],
        [2, "HQ Zone", "Zion Gate BC", "Zion Cell 2", "Mmesoma Ozor", "Behind HQ Fellowship Hall (Saturdays 5:00 PM)"],
        [3, "Franco Zone", "Grace BC", "Franco Cell 1", "Ebuka Aziekwe", "Franco Block A Lounge (Sundays 4:00 PM)"],
        [4, "Hilltop Zone", "Victory BC", "Hilltop Cell 1", "Tony Ubagu", "Hilltop Gate Annex (Wednesdays 5:00 PM)"]
    ]

    for row_idx, row_data in enumerate(sample_cells, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_cells.cell(row=row_idx, column=col_idx, value=val)
            cell.font = normal_font
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")

    # Auto-adjust column widths for all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if val_str.startswith('='):
                    max_len = max(max_len, 15)
                else:
                    max_len = max(max_len, len(val_str))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output_file = "NCF_Database.xlsx"
    wb.save(output_file)
    print(f"✅ NCF Database Excel file successfully generated: {output_file}")

if __name__ == "__main__":
    create_ncf_database()
